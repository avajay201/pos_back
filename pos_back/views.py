from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, serializers
from .models import Order, CustomUser
from datetime import datetime, timedelta
import json
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from urllib.parse import unquote
from django.contrib.auth import authenticate
from django.utils.dateparse import parse_date
from decimal import Decimal
from .apis import API_ENDPOINTS
import requests


class OrderSerializer(serializers.ModelSerializer):
    class Meta:
        model = Order
        fields = "__all__"

    def to_representation(self, instance):
        representation = super().to_representation(instance)
        representation['stored_ids'] = json.loads(instance.stored_ids)
        return representation


class RegisterOrderAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        courses = str(request.data.get('courses'))
        if not courses:
            return Response({'error': 'Courses are required.'}, status=status.HTTP_400_BAD_REQUEST)

        order = Order.objects.create(merchant=request.user, student_name=request.data['student_name'], address=request.data['address'], whatsapp_number=request.data['whatsapp_number'], total_price=request.data.get('totalPrice', 0))
        order.stored_ids = courses.replace("'", '"')
        order.save()
        order_data = {
            'student_name': order.student_name,
            'address': order.address,
            'whatsapp_number': order.whatsapp_number,
            'total_price': order.total_price,
            'courses': json.loads(order.stored_ids),
            'ordered_at': order.ordered_at.strftime("%Y-%m-%d %H:%M:%S")
        }
        return Response(order_data, status=status.HTTP_201_CREATED)


class OrdersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        if not from_date or not to_date:
            return Response({"error": "from_date and to_date are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from_date = datetime.strptime(from_date, "%Y-%m-%d")
            to_date = datetime.strptime(to_date, "%Y-%m-%d")
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        orders = Order.objects.filter(ordered_at__date__range=[from_date, to_date], merchant=request.user)
        serializer = OrderSerializer(orders, many=True)

        return Response({"orders": serializer.data}, status=200)


class LoginAPI(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        if not username or not password:
            return Response({'error': 'Username and password are required'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(username=username, password=password)

        # if user is None or not user.api_token:
        #     response = requests.post(API_ENDPOINTS.get('login'), data={"merchantID": username, "password": password})
        #     if response.status_code != 200:
        #         return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
        #     data = response.json()
        #     user = CustomUser.objects.create_user(username=username, password=password)

        refresh = RefreshToken.for_user(user)

        return Response({
            'access': str(refresh.access_token),
        }, status=status.HTTP_200_OK)


def export_xlsx(request):
    merchant_username = request.GET.get("merchant__username")
    ordered_at_gte = request.GET.get("ordered_at__range__gte")
    ordered_at_lte = request.GET.get("ordered_at__range__lte")

    queryset = Order.objects.all()

    if merchant_username:
        queryset = queryset.filter(merchant__username=unquote(merchant_username))

    if ordered_at_gte:
        ordered_at_gte = parse_date(ordered_at_gte)
        if ordered_at_gte:
            queryset = queryset.filter(ordered_at__date__gte=ordered_at_gte)

    if ordered_at_lte:
        ordered_at_lte = parse_date(ordered_at_lte)
        if ordered_at_lte:
            queryset = queryset.filter(ordered_at__date__lte=ordered_at_lte)

    total_price = sum(order.total_price for order in queryset) or 0.0
    payable_amount = total_price
    if merchant_username and queryset.count() > 0:
        merchant = CustomUser.objects.filter(username=merchant_username).first()
        if merchant:
            payable_amount = total_price - (total_price * (Decimal(merchant.percentage) / Decimal(100)))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders Data"

    headers = ["Merchant", "Student Name", "Total Price", "Whatsapp Number", "Ordered At"]
    worksheet.append(headers)
    
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    for col_num, header in enumerate(headers, 1):
        col_letter = worksheet.cell(row=1, column=col_num).column_letter
        worksheet.cell(row=1, column=col_num, value=header).font = header_font
        worksheet.column_dimensions[col_letter].width = 20

    row_num = 2
    for order in queryset:
        worksheet.append([
            order.merchant.username,
            order.student_name,
            float(order.total_price),
            order.whatsapp_number,
            order.ordered_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

        for col_num in range(1, len(headers) + 1):
            worksheet.cell(row=row_num, column=col_num).alignment = center_alignment

        row_num += 1

    summary_font = Font(bold=True)

    worksheet.append(["Total Price", float(total_price)])
    worksheet.append(["Payable Amount", float(payable_amount)])

    for col_num in range(1, 4):
        worksheet.cell(row=row_num, column=col_num).alignment = center_alignment
        worksheet.cell(row=row_num + 1, column=col_num).alignment = center_alignment
        worksheet.cell(row=row_num, column=col_num).font = summary_font
        worksheet.cell(row=row_num + 1, column=col_num).font = summary_font

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="filtered_orders.xlsx"'
    workbook.save(response)

    return response


class TeacherAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            response = requests.get(API_ENDPOINTS.get('teachers'))
            response.raise_for_status()
            data = response.json()
            teachers = data.get('data', [])
            return Response(teachers, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Failed to fetch teachers."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class TeacherCourseSectionsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, teacher_id):
        try:
            response = requests.get(API_ENDPOINTS.get('teachers_course_sections') + str(teacher_id) + '/')
            response.raise_for_status()
            data = response.json()
            course_sections = data.get('data', {})
            course_sections = course_sections.get('courses', {})
            return Response(course_sections, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Failed to fetch courses."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GradesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            response = requests.get(API_ENDPOINTS.get('grades'))
            response.raise_for_status()
            data = response.json()
            grades = data.get('student_specialities', [])
            return Response(grades, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Failed to fetch grades."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GradeCourseSectionsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        grade_id = request.GET.get("grade_id")
        if not grade_id:
            return Response({"error": "grade_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            response = requests.get(f"{API_ENDPOINTS.get('courses')}/?grade={grade_id}")
            response.raise_for_status()
            data = response.json()
            course_sections = data.get('data', [])
            return Response(course_sections, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Failed to fetch grades."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CourseCouponsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        course_id = request.data.get("course_id")
        if not course_id:
            return Response({"error": "course_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        is_bulk = request.data.get("is_bulk")
        if is_bulk:
            is_bulk = True
        else:
            is_bulk = False

        if is_bulk and not request.user.can_generate_bulk_coupons:
            return Response({"error": "You don't have permission to generate bulk coupons."}, status=status.HTTP_403_FORBIDDEN)

        try:
            if is_bulk:
                response = requests.post(API_ENDPOINTS.get('coupons'), data={"sectionID": course_id}, headers={
                    "Authorization": f"Bearer {request.user.api_token}",
                })
            else:
                response = requests.post(API_ENDPOINTS.get('coupons'), data={"sectionID": course_id})

            response.raise_for_status()
            data = response.json()
            data = data.get('data', {})
            coupon_codes = data.get('coupon_code', {})
            return Response(coupon_codes, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Failed to fetch coupon code."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CoursesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        try:
            response = requests.get(f"{API_ENDPOINTS.get('courses')}{id}")
            response.raise_for_status()
            data = response.json()
            courses = data.get('data', {})
            courses = courses.get('sections', [])
            return Response(courses, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": "Failed to fetch courses."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
